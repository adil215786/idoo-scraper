import os
import time
import resend
import logging
from datetime import datetime
from glob import glob
from dotenv import load_dotenv
import platform
import subprocess
import requests
import json
import traceback

import pandas as pd
from pandas import DataFrame

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException


# Load environment variables
load_dotenv()


class WebhookHandler(logging.Handler):
    """Send critical logs to Discord or Slack webhook"""
    def __init__(self, webhook_url, service_name="T Mobile Scraper"):
        super().__init__()
        self.webhook_url = webhook_url
        self.service_name = service_name

    def emit(self, record):
        if record.levelno >= logging.ERROR:
            try:
                log_entry = self.format(record)

                if "discord.com" in self.webhook_url:
                    payload = {
                        "content": f"ðŸš¨ **{self.service_name} Error**",
                        "embeds": [{
                            "title": "Error Report",
                            "description": f"```{log_entry}```",
                            "color": 15158332,
                            "timestamp": datetime.utcnow().isoformat(),
                            "footer": {"text": platform.node()}
                        }]
                    }
                else:
                    payload = {
                        "text": f"ðŸš¨ *{self.service_name} Error*\n```{log_entry}```"
                    }

                requests.post(self.webhook_url, json=payload, timeout=5)
            except Exception:
                pass


def setup_logging():
    """Setup local and remote logging"""
    handlers = [
        logging.FileHandler('scraper.log'),
        logging.StreamHandler()
    ]

    webhook_url = os.getenv('WEBHOOK_URL')
    if webhook_url:
        handlers.append(WebhookHandler(webhook_url))

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=handlers
    )
    return logging.getLogger(__name__)


logger = setup_logging()
logger.info(f"Scraper starting on {platform.node()}")

root_path = os.getcwd()


def cleanup_chrome_processes():
    """Clean up any hanging Chrome processes"""
    try:
        if platform.system() == "Windows":
            subprocess.run(["taskkill", "/F", "/IM", "chrome.exe"], capture_output=True)
            subprocess.run(["taskkill", "/F", "/IM", "chromedriver.exe"], capture_output=True)
        else:
            subprocess.run(["pkill", "-f", "chrome"], capture_output=True)
            subprocess.run(["pkill", "-f", "chromedriver"], capture_output=True)
        time.sleep(2)
    except Exception:
        pass


def create_download_directory():
    """Create download directory if it doesn't exist"""
    download_dir = os.path.join(root_path, "download_files")
    if not os.path.exists(download_dir):
        os.makedirs(download_dir)
    return os.path.abspath(download_dir)


def wait_for_download(dir_path, target_name="ReOrder Custom Report.xlsx", timeout=180):
    start = time.time()
    final_path = os.path.join(dir_path, target_name)
    while time.time() - start < timeout:
        partials = glob(os.path.join(dir_path, "*.crdownload"))
        if not partials:
            if os.path.exists(final_path):
                return final_path

            xlsxs = sorted(glob(os.path.join(dir_path, "*.xlsx")), key=os.path.getmtime, reverse=True)
            if xlsxs:
                newest = xlsxs[0]
                try:
                    if newest != final_path:
                        os.replace(newest, final_path)
                except PermissionError:
                    pass
                if os.path.exists(final_path):
                    return final_path
        time.sleep(1)
    return None


def driverinitialize(use_proxy=False):
    """
    Initialize Chrome driver.

    Fix applied:
    No hard coded Chrome major version.
    undetected_chromedriver will auto match the installed Chrome version.
    Fallback uses Selenium Manager (Selenium 4.6+) to auto provision driver.
    Detects headless environment (GitHub Actions, Docker, etc.)
    """
    dl_dir = create_download_directory()
    logger.info(f"Downloads will save to: {dl_dir}")

    cleanup_chrome_processes()

    prefs = {
        "download.default_directory": dl_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    }

    # Detect if running in headless environment (GitHub Actions, Docker, etc.)
    is_headless_env = os.getenv('CI') == 'true' or os.getenv('GITHUB_ACTIONS') == 'true' or not os.environ.get('DISPLAY')
    
    if is_headless_env:
        logger.info("Headless environment detected (GitHub Actions/CI) - using headless mode")

    try:
        import undetected_chromedriver as uc

        chrome_options = uc.ChromeOptions()
        if is_headless_env:
            chrome_options.add_argument("--headless=new")
            chrome_options.add_argument("--window-size=1920,1080")
        else:
            chrome_options.add_argument("--start-maximized")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_experimental_option("prefs", prefs)

        driver = uc.Chrome(options=chrome_options)

        driver.implicitly_wait(10)
        logger.info("Undetected Chrome driver initialized successfully")
        return driver

    except Exception as e:
        logger.error(f"Undetected Chrome init failed, falling back to Selenium Chrome. Error: {e}")

        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options

            chrome_options = Options()
            if is_headless_env:
                chrome_options.add_argument("--headless=new")
                chrome_options.add_argument("--window-size=1920,1080")
            else:
                chrome_options.add_argument("--start-maximized")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            chrome_options.add_experimental_option("prefs", prefs)
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option("useAutomationExtension", False)

            driver = webdriver.Chrome(options=chrome_options)

            driver.implicitly_wait(10)
            logger.info("Selenium Chrome driver initialized successfully")
            return driver

        except Exception as e2:
            logger.error(f"Failed to initialize Selenium Chrome driver: {e2}")
            raise


def wait_for_element(driver, xpath, timeout=10, condition=EC.presence_of_element_located):
    """Wait for element with proper error handling"""
    try:
        element = WebDriverWait(driver, timeout).until(
            condition((By.XPATH, xpath))
        )
        return element
    except TimeoutException:
        logger.warning(f"Element not found within {timeout} seconds: {xpath}")
        return None
    except Exception as e:
        logger.error(f"Error waiting for element {xpath}: {e}")
        return None


def do_login(driver, user_id, password, max_retries=1):
    """Login with improved error handling and retry logic"""
    for attempt in range(max_retries):
        try:
            logger.info(f"Login attempt {attempt + 1}/{max_retries}")

            driver.get("https://www.t-mobiledealerordering.com/b2b_tmo/init.do")
            time.sleep(3)

            userid_field = wait_for_element(driver, '//input[@id="userid"]', timeout=15)
            if not userid_field:
                logger.error("Login form not found")
                continue

            userid_field.clear()
            time.sleep(0.5)
            userid_field.send_keys(user_id)

            password_field = wait_for_element(driver, '//input[@id="password"]')
            if password_field:
                password_field.clear()
                time.sleep(0.5)
                password_field.send_keys(password)

            agree_checkbox = wait_for_element(driver, '//input[@name="AgreeTerms"]')
            if agree_checkbox:
                agree_checkbox.click()
                time.sleep(1)

            login_button = wait_for_element(driver, '//a[@name="login"]')
            if login_button:
                login_button.click()
                time.sleep(5)

            for _ in range(15):
                try:
                    driver.find_element(By.XPATH, '//frameset[@id="isaTopFS"]')
                    logger.info("Login successful!")
                    return True
                except NoSuchElementException:
                    time.sleep(1)

            logger.warning(f"Login attempt {attempt + 1} failed")

            if attempt < max_retries - 1:
                logger.info("Refreshing page and retrying...")
                driver.refresh()
                time.sleep(5)

        except Exception as e:
            logger.error(f"Error during login attempt {attempt + 1}: {e}")
            if attempt < max_retries - 1:
                time.sleep(5)

    logger.error("All login attempts failed")
    return False


def download_report(report_user_id, report_password):
    """Download report with improved error handling"""
    report_driver = None
    try:
        report_driver = driverinitialize()

        report_driver.set_page_load_timeout(300)
        report_driver.implicitly_wait(30)

        login_success = False
        for attempt in range(3):
            try:
                report_driver.get("https://www.myrtpos.com/newbdi/index.fwx")
                time.sleep(5)

                userid_field = wait_for_element(report_driver, '//input[@name="secUserID"]')
                if userid_field:
                    userid_field.clear()
                    userid_field.send_keys(report_user_id)

                password_field = wait_for_element(report_driver, '//input[@name="secPassword"]')
                if password_field:
                    password_field.clear()
                    password_field.send_keys(report_password)

                login_button = wait_for_element(report_driver, '//input[@value="Login"]')
                if login_button:
                    login_button.click()
                    time.sleep(5)

                time.sleep(3)

                inventory_menu = None
                try:
                    inventory_menu = report_driver.find_element(By.XPATH, '//span[contains(.,"Inventory Report") and img[contains(@src,"cellularphone.png")]]')
                except Exception:
                    try:
                        inventory_menu = report_driver.find_element(By.XPATH, '//span[contains(.,"Inventory Report")]')
                    except Exception:
                        try:
                            inventory_menu = report_driver.find_element(By.XPATH, '//*[span[contains(.,"Inventory Report")]]')
                        except Exception:
                            inventory_menu = None

                if not inventory_menu:
                    raise Exception("Inventory Report menu not found")

                login_success = True
                logger.info("RT POS login successful")

                logger.info("Clicking Inventory Report menu...")
                try:
                    parent = inventory_menu.find_element(By.XPATH, '..')
                    parent.click()
                except Exception:
                    try:
                        inventory_menu.click()
                    except Exception:
                        report_driver.execute_script("arguments[0].click();", inventory_menu)

                time.sleep(3)

                reorder_link = None
                try:
                    reorder_link = report_driver.find_element(By.XPATH, '//b[contains(text(),"Re-Order Report By Store")]')
                except Exception:
                    try:
                        reorder_link = report_driver.find_element(By.XPATH, '//a[b[contains(text(),"Re-Order Report By Store")]]')
                    except Exception:
                        try:
                            reorder_link = report_driver.find_element(By.XPATH, '//*[contains(.,"Re-Order Report By Store")]')
                        except Exception:
                            reorder_link = None

                if reorder_link:
                    logger.info("Clicking Re Order Report By Store...")
                    try:
                        reorder_link.click()
                    except Exception:
                        try:
                            parent = reorder_link.find_element(By.XPATH, '..')
                            parent.click()
                        except Exception:
                            report_driver.execute_script("arguments[0].click();", reorder_link)
                    time.sleep(5)
                else:
                    logger.info("Re Order link not found, navigating directly...")
                    report_driver.get("https://www.myrtpos.com/newbdi/Reorder_Custom.fwx")
                    time.sleep(5)

                break

            except Exception as e:
                if attempt < 2:
                    logger.warning(f"RT POS login attempt {attempt + 1} failed: {e}")
                    screenshot_path = os.path.join(create_download_directory(), f"rtpos_error_{report_user_id}_{attempt}.png")
                    try:
                        report_driver.save_screenshot(screenshot_path)
                        logger.info(f"Error screenshot saved to: {screenshot_path}")
                    except Exception:
                        pass
                    continue
                logger.error(f"RT POS login failed after 3 attempts: {e}")

        if not login_success:
            logger.error("Failed to login to RT POS after 3 attempts")
            return False

        time.sleep(5)

        days_field = wait_for_element(report_driver, '//input[@name="frmDays"]')
        if days_field:
            days_field.click()
            time.sleep(0.5)
            days_field.clear()
            time.sleep(1)
            days_field.send_keys("7")
            time.sleep(1)
            logger.info("Set days to 7")

        logger.info("Looking for Generate button...")
        generate_button = wait_for_element(report_driver, '//span[contains(text(),"Generate")]')

        if not generate_button:
            logger.error("Generate button not found")
            return False

        logger.info("Clicking Generate button...")
        clicked = False
        try:
            generate_button.click()
            clicked = True
        except Exception:
            try:
                report_driver.execute_script("arguments[0].click();", generate_button)
                clicked = True
            except Exception:
                clicked = False

        if not clicked:
            logger.error("Failed to click Generate button")
            return False

        logger.info("Form submitted, waiting for report generation...")
        time.sleep(5)

        max_wait_time = 300
        poll_interval = 5
        elapsed_time = 0

        while elapsed_time < max_wait_time:
            try:
                export_button = report_driver.find_element(By.XPATH, '//div/i[@class="dx-icon dx-icon-export-excel-button"]')
                if export_button:
                    logger.info(f"Report generated successfully after {elapsed_time} seconds")

                    logger.info("Clicking Excel export button...")
                    export_button.click()

                    logger.info("Waiting for download to complete...")
                    time.sleep(5)

                    logger.info("Report downloaded successfully")
                    return True

            except NoSuchElementException:
                pass
            except Exception as e:
                logger.debug(f"Polling check at {elapsed_time}s: {type(e).__name__}")

            try:
                error_elements = report_driver.find_elements(By.XPATH, '//*[contains(text(),"Error") or contains(text(),"error") or contains(text(),"failed")]')
                for elem in error_elements:
                    if elem.is_displayed() and elem.text:
                        logger.error(f"Error message found: {elem.text}")
                        return False
            except Exception:
                pass

            time.sleep(poll_interval)
            elapsed_time += poll_interval

            if elapsed_time % 30 == 0:
                logger.info(f"Still waiting for report... ({elapsed_time} seconds elapsed)")

        logger.error(f"Report generation timed out after {max_wait_time} seconds")
        try:
            report_driver.save_screenshot(f"report_timeout_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
        except Exception:
            pass
        return False

    except Exception as e:
        logger.error(f"Error downloading report: {e}")
        return False
    finally:
        if report_driver:
            try:
                report_driver.quit()
                logger.info("Report browser closed")
            except Exception as e:
                logger.error(f"Error closing report driver: {e}")


def create_new_report(ids, stock_data_rows, subject, output_file):
    """Create new report with enhanced formatting"""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        download_dir = create_download_directory()
        file_path = os.path.join(download_dir, "ReOrder Custom Report.xlsx")

        if not os.path.exists(file_path):
            logger.error("Report file not found")
            return False

        df = pd.read_excel(file_path)
        market = ""
        store_id = ""
        store_name = ""
        datarows = []

        has_on_hand = 'On Hand' in df.columns
        has_on_po = 'On PO' in df.columns

        for _, row in df.iterrows():
            data_row = row.to_dict()
            item_number = str(data_row.get("Item Number"))

            if item_number == "nan":
                manufacturer = str(data_row.get("Manufacturer"))
                if manufacturer == "nan":
                    continue
                if "Market" in manufacturer:
                    market = manufacturer.replace("Market:", "").strip()
                elif "StoreID" in manufacturer:
                    store_id = manufacturer.replace("StoreID:", "").strip()
                elif "Store Name" in manufacturer:
                    store_name = manufacturer.replace("Store Name:", "").strip()
                continue

            if item_number not in ids:
                continue

            on_hand = data_row.get("On Hand") if has_on_hand else ""
            on_po = data_row.get("On PO") if has_on_po else ""

            datarow = [
                market, store_id, store_name,
                data_row.get("Manufacturer"), item_number,
                data_row.get("Item Description"),
                on_hand, on_po,
                data_row.get("7 Days"), data_row.get("Item Cost"),
                data_row.get("Total Qty"), data_row.get("Suggested")
            ]
            datarows.append(datarow)

        if not datarows:
            logger.warning("No matching items found in report")
            return False

        out_df = DataFrame(datarows, columns=[
            'Market', 'StoreID', 'Store Name', 'Manufacturer',
            'Item Number', 'Item Description', 'On Hand', 'On PO',
            '7 Days', 'Item Cost', 'Total Qty', 'Suggested'
        ])
        out_df.drop_duplicates(inplace=True)

        formatted_df = out_df.copy()
        formatted_df = formatted_df.drop(columns=['StoreID', 'Manufacturer'])

        cols = formatted_df.columns.tolist()
        item_cost_col = cols.pop(7)
        cols.insert(0, item_cost_col)
        formatted_df = formatted_df[cols]

        formatted_df.columns = [
            'Item Cost', 'Market', 'Store Name', 'Item Number', 'Item Description',
            'On Hand', 'On PO', '7 Days', 'Total Qty', 'Suggested'
        ]

        formatted_df['Qty'] = 0
        formatted_df['0'] = 0
        formatted_df['Shipping'] = ''
        formatted_df['Total'] = ''
        formatted_df['Your Total'] = ''
        formatted_df['Difference'] = ''

        stock_df = pd.DataFrame(stock_data_rows, columns=['SKU', 'Quantity'])

        output_path = os.path.join(download_dir, output_file)
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            out_df.to_excel(writer, sheet_name="report", index=False)
            formatted_df.to_excel(writer, sheet_name="Phone distribution idoo", index=False)
            stock_df.to_excel(writer, sheet_name="stock_quantity", index=False)

            workbook = writer.book
            worksheet = workbook["Phone distribution idoo"]

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            center_alignment = Alignment(horizontal='center', vertical='center')

            orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

            red_font = Font(color='FF0000')

            max_row = worksheet.max_row

            worksheet.column_dimensions['A'].width = 10.00
            worksheet.column_dimensions['B'].width = 10.00
            worksheet.column_dimensions['C'].width = 20.00
            worksheet.column_dimensions['D'].width = 15.00
            worksheet.column_dimensions['E'].width = 30.00
            worksheet.column_dimensions['F'].width = 10.00
            worksheet.column_dimensions['G'].width = 10.00
            worksheet.column_dimensions['H'].width = 10.00
            worksheet.column_dimensions['I'].width = 10.00
            worksheet.column_dimensions['J'].width = 10.00
            worksheet.column_dimensions['K'].width = 10.00
            worksheet.column_dimensions['L'].width = 10.00
            worksheet.column_dimensions['M'].width = 10.00
            worksheet.column_dimensions['N'].width = 10.00
            worksheet.column_dimensions['O'].width = 10.00
            worksheet.column_dimensions['P'].width = 10.00

            current_store = None
            use_blue = True

            for row in range(1, max_row + 1):
                for col in range(1, 17):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = center_alignment

                    if col in [15, 16]:
                        cell.fill = orange_fill

                    if col == 16 and row > 1:
                        cell.font = red_font
                        cell.value = f'=O{row}-N{row}'

                    if col == 3 and row > 1:
                        store_val = cell.value
                        if store_val != current_store:
                            current_store = store_val
                            use_blue = not use_blue

                        for row_col in range(1, 17):
                            if row_col not in [15, 16]:
                                worksheet.cell(row=row, column=row_col).fill = light_blue_fill if use_blue else white_fill

            for row in range(2, max_row + 1):
                worksheet[f'L{row}'].value = f'=K{row}*A{row}'

            worksheet['L1'].value = f'=SUM(L2:L{max_row})'

            store_groups = []
            current_store = None
            start_row = 2

            for row in range(2, max_row + 1):
                store_val = worksheet[f'C{row}'].value
                if store_val != current_store:
                    if current_store is not None:
                        store_groups.append((current_store, start_row, row - 1))
                    current_store = store_val
                    start_row = row

            if current_store is not None:
                store_groups.append((current_store, start_row, max_row))

            for _, start_row, end_row in store_groups:
                if start_row < end_row:
                    worksheet.merge_cells(f'N{start_row}:N{end_row}')
                    worksheet.merge_cells(f'O{start_row}:O{end_row}')
                    worksheet.merge_cells(f'P{start_row}:P{end_row}')

                worksheet[f'N{start_row}'].value = f'=SUM(L{start_row}:L{end_row})'
                worksheet[f'N{start_row}'].alignment = center_alignment
                worksheet[f'N{start_row}'].border = thin_border

                worksheet[f'O{start_row}'].alignment = center_alignment
                worksheet[f'O{start_row}'].border = thin_border
                worksheet[f'O{start_row}'].fill = orange_fill

                worksheet[f'P{start_row}'].value = f'=O{start_row}-N{start_row}'
                worksheet[f'P{start_row}'].alignment = center_alignment
                worksheet[f'P{start_row}'].border = thin_border
                worksheet[f'P{start_row}'].font = red_font
                worksheet[f'P{start_row}'].fill = orange_fill

            for col in range(1, 17):
                header_cell = worksheet.cell(row=1, column=col)
                header_cell.font = Font(bold=True)
                header_cell.border = thin_border
                header_cell.alignment = center_alignment
                if col not in [15, 16]:
                    header_cell.fill = white_fill

        logger.info("Enhanced Excel file created successfully")

        try:
            os.remove(file_path)
        except Exception as e:
            logger.error(f"Error removing original file: {e}")

        return True

    except Exception as e:
        logger.error(f"Error creating report: {e}")
        logger.error(traceback.format_exc())
        return False


def safe_quit(driver):
    """Safely quit the driver"""
    try:
        driver.quit()
        time.sleep(2)
    except Exception:
        pass


def main():
    """Main function with comprehensive error handling"""
    driver = None
    total_start_time = time.time()

    try:
        cred_file = "cred.txt"
        if not os.path.exists(cred_file):
            logger.error(f"Credentials file {cred_file} not found")
            return

        with open(cred_file, "r") as f:
            creds = f.read().split("\n")

        driver = driverinitialize()

        for cred in creds:
            if not cred.strip():
                continue

            try:
                user_part, report_part = cred.split("||")
                user_id, password = user_part.split("|")
                report_user_id, report_password = report_part.split("|")
            except ValueError as e:
                logger.error(f"Invalid credential format: {e}")
                continue

            logger.info(f"Processing user: {user_id}")

            now = datetime.now()
            today_date = now.strftime('%m-%d-%Y')
            account_label = user_id.upper() if user_id.lower().startswith('iot') else user_id
            subject = f"INVENTORY - {account_label} - {today_date}"
            output_file = f"IDOO-{account_label}-{today_date}.xlsx"

            datarows = []
            stocks_data_rows = []

            driver.get("https://www.t-mobiledealerordering.com/b2b_tmo/init.do")
            time.sleep(3)

            if not do_login(driver, user_id, password):
                logger.error(f"Login failed for user {user_id}")
                try:
                    driver.save_screenshot(os.path.join(root_path, f"error_screenshot_{user_id}.png"))
                except Exception:
                    pass
                continue

            frame_retries = 0
            while frame_retries < 3:
                try:
                    driver.switch_to.default_content()
                    driver.switch_to.frame("isaTop")
                    time.sleep(1)
                    driver.switch_to.frame("header")
                    time.sleep(1)
                    break
                except Exception as e:
                    frame_retries += 1
                    logger.warning(f"Frame navigation attempt {frame_retries} failed: {e}")
                    if frame_retries < 3:
                        driver.refresh()
                        time.sleep(5)
                    else:
                        logger.error("Failed to navigate frames after 3 attempts")
                        continue

            catalog_clicked = False
            for attempt in range(3):
                try:
                    catalog_button = driver.find_element(By.XPATH, '//a[@onclick="show_catalog_view()"]')
                    catalog_button.click()
                    catalog_clicked = True
                    logger.info("Clicked catalog view button")
                    break
                except Exception as e:
                    logger.warning(f"Catalog button attempt {attempt + 1} failed: {e}")
                    if attempt < 2:
                        driver.get(driver.current_url)
                        time.sleep(10)
                        driver.switch_to.default_content()
                        driver.switch_to.frame("isaTop")
                        time.sleep(1)
                        driver.switch_to.frame("header")
                        time.sleep(1)

            if not catalog_clicked:
                logger.error("Failed to click catalog view button")
                continue

            time.sleep(10)

            driver.switch_to.default_content()
            form_frame_found = False
            for attempt in range(3):
                try:
                    driver.switch_to.frame("isaTop")
                    time.sleep(1)
                    driver.switch_to.frame("form_input")
                    form_frame_found = True
                    logger.info("Successfully navigated to form_input frame")
                    break
                except Exception as e:
                    logger.warning(f"Form frame attempt {attempt + 1} failed: {e}")
                    if attempt < 2:
                        try:
                            driver.switch_to.default_content()
                            driver.switch_to.frame("isaTop")
                            driver.switch_to.frame("header")
                            driver.find_element(By.XPATH, '//a[@onclick="show_catalog_view()"]').click()
                            time.sleep(5)
                        except Exception:
                            driver.get(driver.current_url)
                            time.sleep(10)

            if not form_frame_found:
                logger.error("Failed to navigate to form_input frame")
                continue

            nodes = driver.find_elements(By.XPATH, '//div[contains(@class,"catItemList-holder")]/div[@class="catalauge-item-holder "]')

            for node in nodes:
                try:
                    node_sku = node.find_element(By.XPATH, './/div[@class="cat-prd-id"]').get_attribute("innerText").strip()
                    qty_text = node.find_element(By.XPATH, './/td[@class="cat-prd-qty"]').get_attribute("innerText").strip()
                    node_allocation_available_qty = qty_text.replace("Allocation :", "").strip().split("of")[0].strip()

                    if int(node_allocation_available_qty) > 0:
                        logger.info(f"Stock added for SKU: {node_sku}")
                        datarows.append(node_sku)
                        stocks_data_rows.append([node_sku, node_allocation_available_qty])
                except Exception as e:
                    logger.error(f"Error processing node: {e}")
                    continue

            if datarows:
                filter_button = wait_for_element(driver, '//input[@id="filterAllocBtn"]')
                if filter_button:
                    filter_button.click()
                    time.sleep(10)
                    try:
                        driver.save_screenshot(os.path.join(create_download_directory(), "phone_screenshot.png"))
                    except Exception:
                        pass
                    time.sleep(3)

            cpo_link = wait_for_element(driver, '//div[@class="cat-secnav-areaname"]/a/span[contains(text(),"CPO")]')
            if cpo_link:
                cpo_link.click()
                time.sleep(8)

                cpo_nodes = driver.find_elements(By.XPATH, '//div[contains(@class,"catItemList-holder")]/div[@class="catalauge-item-holder "]')

                for node in cpo_nodes:
                    try:
                        node_sku = node.find_element(By.XPATH, './/div[@class="cat-prd-id"]').get_attribute("innerText").strip()
                        qty_text = node.find_element(By.XPATH, './/td[@class="cat-prd-qty"]').get_attribute("innerText").strip()
                        node_allocation_available_qty = qty_text.replace("Allocation :", "").strip().split("of")[0].strip()

                        if int(node_allocation_available_qty) > 0:
                            logger.info(f"CPO Stock added for SKU: {node_sku}")
                            datarows.append(node_sku)
                            stocks_data_rows.append([node_sku, node_allocation_available_qty])
                    except Exception as e:
                        logger.error(f"Error processing CPO node: {e}")
                        continue

                if cpo_nodes:
                    try:
                        driver.save_screenshot(os.path.join(create_download_directory(), "cpo_screenshot.png"))
                    except Exception:
                        pass
                    time.sleep(3)

            SIM_CARD_SKU = "NC128TRIPLESIM"
            if SIM_CARD_SKU not in datarows:
                datarows.append(SIM_CARD_SKU)

            if not datarows:
                logger.info("No products have stock available.")
            else:
                logger.info(f"Found {len(datarows)} items with stock")

                if download_report(report_user_id, report_password):
                    if create_new_report(datarows, stocks_data_rows, subject, output_file):
                        logger.info("Process completed successfully")
                    else:
                        logger.error("Failed to create report")
                else:
                    logger.error("Failed to download report")

            logger.info(f"Completed processing for user: {user_id}")

    except Exception as e:
        logger.error(f"Unexpected error in main: {e}")
        logger.error(traceback.format_exc())
        if driver:
            try:
                driver.save_screenshot(os.path.join(root_path, "error_screenshot.png"))
            except Exception:
                pass
    finally:
        if driver:
            try:
                driver.quit()
                driver = None
            except Exception:
                pass

        logger.info("All users processed.")
        total_time = time.time() - total_start_time
        logger.info(f"Total execution time: {total_time:.2f} seconds")


if __name__ == "__main__":
    import warnings
    import sys

    warnings.filterwarnings("ignore")

    if platform.system() == "Windows":
        sys.stderr = open(os.devnull, 'w')
    else:
        sys.stderr = open('/dev/null', 'w')

    try:
        main()
    except KeyboardInterrupt:
        print("\nScript interrupted by user")
    except Exception as e:
        sys.stderr = sys.__stderr__
        logger.error(f"Fatal error: {e}")
    finally:
        sys.exit(0)
